import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# 샘플 데이터 생성
data = {
    'TC ID': ['TC-001', 'TC-002', 'TC-003', 'TC-004', 'TC-005'],
    'Depth 1': ['로그인', '로그인', '회원가입', '프로필', '설정'],
    'Depth 2': ['정상', '비정상', '신규', '수정', '알림'],
    'Steps': [
        '1. 앱을 실행한다\n2. 이메일 입력란에 올바른 이메일을 입력한다\n3. 비밀번호 입력란에 올바른 비밀번호를 입력한다\n4. 로그인 버튼을 클릭한다',
        '1. 앱을 실행한다\n2. 이메일 입력란에 잘못된 형식의 이메일을 입력한다\n3. 비밀번호 입력란에 비밀번호를 입력한다\n4. 로그인 버튼을 클릭한다',
        '1. 회원가입 화면으로 이동한다\n2. 필수 항목(이메일, 비밀번호, 이름)을 모두 입력한다\n3. 이용약관에 동의한다\n4. 가입하기 버튼을 클릭한다',
        '1. 프로필 화면으로 이동한다\n2. 프로필 사진 변경 버튼을 클릭한다\n3. 갤러리에서 사진을 선택한다\n4. 저장 버튼을 클릭한다',
        '1. 설정 화면으로 이동한다\n2. 알림 설정 메뉴를 클릭한다\n3. 푸시 알림 토글을 ON으로 변경한다\n4. 저장 버튼을 클릭한다'
    ],
    'Expected Result': [
        '로그인이 성공하고 메인 화면으로 이동한다',
        '이메일 형식이 올바르지 않다는 오류 메시지가 표시된다',
        '회원가입이 완료되고 환영 메시지가 표시된다',
        '프로필 사진이 성공적으로 변경되고 저장 완료 메시지가 표시된다',
        '푸시 알림 설정이 저장되고 확인 메시지가 표시된다'
    ]
}

# DataFrame 생성
df = pd.DataFrame(data)

# Workbook 생성
wb = Workbook()
ws = wb.active
ws.title = "Test Cases"

# DataFrame을 worksheet에 추가
for r in dataframe_to_rows(df, index=False, header=True):
    ws.append(r)

# 스타일 설정
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
border = Border(
    left=Side(style='thin', color='000000'),
    right=Side(style='thin', color='000000'),
    top=Side(style='thin', color='000000'),
    bottom=Side(style='thin', color='000000')
)

# 헤더 스타일 적용
for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = border

# 데이터 셀 스타일 적용
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    for cell in row:
        cell.border = border
        cell.alignment = Alignment(vertical='top', wrap_text=True)

# 열 너비 조정
ws.column_dimensions['A'].width = 12  # TC ID
ws.column_dimensions['B'].width = 15  # Depth 1
ws.column_dimensions['C'].width = 15  # Depth 2
ws.column_dimensions['D'].width = 50  # Steps
ws.column_dimensions['E'].width = 50  # Expected Result

# 행 높이 자동 조정
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    ws.row_dimensions[row[0].row].height = 80

# 파일 저장
import datetime
timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
output_file = f'한국어 테스트케이스_{timestamp}.xlsx'
wb.save(output_file)
print(f"✓ 샘플 파일이 생성되었습니다: {output_file}")
