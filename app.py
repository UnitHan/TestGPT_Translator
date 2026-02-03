# -*- coding: utf-8 -*-
import sys
import io
import logging
from logging.handlers import RotatingFileHandler
import os
import time
from datetime import datetime
from pathlib import Path
from threading import Thread, Lock
from uuid import uuid4
import webbrowser
from urllib.request import urlopen
from urllib.error import URLError
import socket
import psutil

def get_console_stream():
    for stream in (sys.stdout, sys.__stdout__, sys.stderr, sys.__stderr__):
        if stream and hasattr(stream, "write"):
            return stream
    return None


# Windows 콘솔에서 UTF-8 출력 지원
if sys.platform == "win32":
    if sys.stdout and hasattr(sys.stdout, "buffer"):
        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    if sys.stderr and hasattr(sys.stderr, "buffer"):
        sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

def get_log_dir() -> Path:
    override = os.environ.get("TRANSLATOR_LOG_DIR")
    if override:
        return Path(override).expanduser().resolve()
    if sys.platform == "win32":
        return Path("C:/translation_log")
    return Path.home() / "translation_log"


LOG_DIR = get_log_dir()
LOG_FILE = LOG_DIR / "translation-server.log"

# 로깅 설정
console_stream = get_console_stream()
log_handlers = [logging.StreamHandler(console_stream)] if console_stream else []
try:
    LOG_DIR.mkdir(parents=True, exist_ok=True)
    file_handler = RotatingFileHandler(
        str(LOG_FILE), maxBytes=5 * 1024 * 1024, backupCount=5, encoding="utf-8"
    )
    log_handlers.append(file_handler)
except Exception as log_exc:
    log_handlers.append(logging.StreamHandler(sys.stdout))
    print(f"[Warning] Failed to initialize file logging: {log_exc}")

logging.basicConfig(
    level=logging.INFO,  # DEBUG -> INFO로 변경 (로그 줄이기)
    format='[%(asctime)s] [%(levelname)s] %(message)s',
    handlers=log_handlers,
    force=True
)
logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)

logger.info(f"Log File: {LOG_FILE}")

from flask import Flask, render_template, request, jsonify, send_file, send_from_directory
from flask_cors import CORS
from werkzeug.utils import secure_filename

# 무거운 라이브러리는 필요할 때만 임포트 (지연 임포트)
# from openpyxl import load_workbook
# import google.generativeai as genai

APP_DIR_NAME = "TestGPT-TC-Translator"
DEFAULT_PORT = 5000
MAX_UPLOAD_MB = int(os.environ.get("MAX_UPLOAD_MB", "50"))
MAX_CONTENT_LENGTH = MAX_UPLOAD_MB * 1024 * 1024
ALLOWED_EXTENSIONS = {".xlsx"}
GEMINI_MAX_RETRIES = int(os.environ.get("GEMINI_MAX_RETRIES", "3"))
GEMINI_RETRY_BACKOFF = float(os.environ.get("GEMINI_RETRY_BACKOFF", "1.0"))


def get_resource_dir() -> Path:
    """Locate bundled resources (templates/icons) for dev or PyInstaller builds."""
    if getattr(sys, "frozen", False):
        # PyInstaller onedir: resources are in _internal folder
        if hasattr(sys, '_MEIPASS'):
            return Path(sys._MEIPASS)
        # Fallback to exe parent directory
        return Path(sys.executable).resolve().parent
    return Path(__file__).resolve().parent


def get_app_data_dir() -> Path:
    """Return a user-writable data directory for uploads/outputs."""
    override = os.environ.get("TRANSLATOR_DATA_DIR")
    if override:
        return Path(override).expanduser().resolve()

    home = Path.home()
    if sys.platform == "win32":
        base = os.environ.get("LOCALAPPDATA") or os.environ.get("APPDATA") or str(home)
    elif sys.platform == "darwin":
        base = str(home / "Library" / "Application Support")
    else:
        base = os.environ.get("XDG_DATA_HOME") or str(home / ".local" / "share")

    return Path(base) / APP_DIR_NAME


RESOURCE_DIR = get_resource_dir()
DATA_DIR = get_app_data_dir()
UPLOAD_FOLDER = DATA_DIR / "uploads"
OUTPUT_FOLDER = DATA_DIR / "outputs"
UPLOAD_FOLDER.mkdir(parents=True, exist_ok=True)
OUTPUT_FOLDER.mkdir(parents=True, exist_ok=True)

app = Flask(
    __name__,
    static_folder=str(RESOURCE_DIR),
    template_folder=str(RESOURCE_DIR / "templates"),
    static_url_path="/static",
)
app.config["MAX_CONTENT_LENGTH"] = MAX_CONTENT_LENGTH
app.json.ensure_ascii = False
CORS(app)

logger.info(f"Resource Directory: {RESOURCE_DIR}")
logger.info(f"Data Directory: {DATA_DIR}")
logger.info(f"Upload Folder: {UPLOAD_FOLDER}")
logger.info(f"Output Folder: {OUTPUT_FOLDER}")

# 번역 상태를 저장할 전역 변수
status_lock = Lock()


def _new_status() -> dict:
    return {
        "job_id": None,
        "progress": 0,
        "total": 0,
        "current": 0,
        "status": "idle",
        "estimated_time": 0,
        "output_file": None,
        "error": None,
        "started_at": None,
        "completed_at": None,
    }


translation_status = _new_status()

# Heartbeat 관리 (브라우저 연결 모니터링)
last_heartbeat = None
heartbeat_lock = Lock()
HEARTBEAT_TIMEOUT = 30  # 초 (30초 동안 heartbeat 없으면 종료)
shutdown_flag = False


def get_status_snapshot() -> dict:
    with status_lock:
        return dict(translation_status)


def update_status(**kwargs) -> None:
    with status_lock:
        translation_status.update(kwargs)


def set_error_once(message: str) -> None:
    with status_lock:
        if not translation_status.get("error"):
            translation_status["error"] = message


def reset_status(job_id: str | None = None, output_file: str | None = None) -> None:
    base = _new_status()
    base["job_id"] = job_id
    base["output_file"] = output_file
    with status_lock:
        translation_status.clear()
        translation_status.update(base)


def is_processing() -> bool:
    with status_lock:
        return translation_status.get("status") == "processing"


def update_heartbeat():
    """브라우저 heartbeat 업데이트"""
    global last_heartbeat
    with heartbeat_lock:
        last_heartbeat = time.time()


def check_heartbeat_timeout():
    """Heartbeat 타임아웃 체크 (백그라운드 스레드)"""
    global shutdown_flag
    while not shutdown_flag:
        time.sleep(5)  # 5초마다 체크
        
        with heartbeat_lock:
            if last_heartbeat is None:
                continue  # 아직 첫 heartbeat 받지 못함
            
            elapsed = time.time() - last_heartbeat
            if elapsed > HEARTBEAT_TIMEOUT:
                logger.warning(f"No heartbeat for {elapsed:.1f} seconds")
                logger.info("Browser appears to be closed. Shutting down server...")
                shutdown_flag = True
                
                # 서버 종료
                import signal
                os.kill(os.getpid(), signal.SIGTERM)
                break


# Gemini API 설정 (환경변수에서 API 키 가져오기) - 지연 로드
gemini_model = None
gemini_model_error = None


def get_gemini_model():
    """Gemini 모델을 필요할 때만 로드"""
    global gemini_model, gemini_model_error
    if gemini_model is not None:
        return gemini_model
    if gemini_model_error:
        return None

    try:
        import google.generativeai as genai

        api_key = os.environ.get("GEMINI_API_KEY")
        if not api_key:
            gemini_model_error = "GEMINI_API_KEY not set"
            logger.warning("[Warning] GEMINI_API_KEY not set")
            return None

        genai.configure(api_key=api_key)
        gemini_model = genai.GenerativeModel("gemini-2.5-flash")
        logger.info("Gemini API initialized with model: gemini-2.5-flash")
    except Exception as e:
        gemini_model_error = str(e)
        logger.error(f"Failed to initialize Gemini API: {e}")

    return gemini_model


def is_allowed_file(filename: str) -> bool:
    return Path(filename).suffix.lower() in ALLOWED_EXTENSIONS


def is_retryable_exception(exc: Exception) -> bool:
    try:
        from google.api_core import exceptions as gexc

        if isinstance(
            exc,
            (
                gexc.DeadlineExceeded,
                gexc.InternalServerError,
                gexc.ResourceExhausted,
                gexc.ServiceUnavailable,
            ),
        ):
            return True
    except Exception:
        pass

    message = str(exc).lower()
    retry_terms = (
        "rate limit",
        "quota",
        "timeout",
        "temporarily",
        "unavailable",
        "internal",
        "429",
        "503",
    )
    return any(term in message for term in retry_terms)


def safe_unlink(path: Path) -> None:
    try:
        path.unlink(missing_ok=True)
    except Exception as exc:
        logger.warning(f"Failed to delete file: {path} ({exc})")

def translate_with_llm(text, context=""):
    """LLM을 사용하여 테스트 케이스를 전문적으로 번역"""
    if not text or not isinstance(text, str) or not text.strip():
        return text

    model = get_gemini_model()  # 필요할 때 로드
    if not model:
        raise RuntimeError("Gemini API model not initialized - check API key")
    
    prompt = f"""You are a senior QA engineer with 30 years of experience in software testing and mobile app testing. 
You are an expert in translating test cases from Korean to English while maintaining technical accuracy and clarity.

Translate the following Korean test case text to English. Keep the translation:
- Professional and technically accurate
- Clear and concise
- Using proper QA/testing terminology
- Maintaining the original meaning and intent
- Preserving line breaks and formatting

{f'Context: {context}' if context else ''}

Korean text to translate:
{text}

Provide ONLY the English translation without any additional explanation or comments."""

    last_error = None
    for attempt in range(1, GEMINI_MAX_RETRIES + 1):
        try:
            logger.debug(f"Translating text ({len(text)} chars, {len(text.split())} words)")
            response = model.generate_content(prompt)

            translated = response.text.strip() if response and hasattr(response, "text") else ""
            if not translated:
                raise RuntimeError("API returned empty or invalid response")

            logger.debug(f"Translation completed ({len(translated)} chars)")
            return translated
        except Exception as exc:
            last_error = exc
            if is_retryable_exception(exc) and attempt < GEMINI_MAX_RETRIES:
                backoff = GEMINI_RETRY_BACKOFF * (2 ** (attempt - 1))
                logger.warning(
                    f"Retryable translation error (attempt {attempt}/{GEMINI_MAX_RETRIES}): {exc}"
                )
                time.sleep(backoff)
                continue
            raise

    raise last_error

def process_excel_translation(input_file: Path, output_file: Path, job_id: str) -> None:
    """엑셀 파일을 읽어 Steps와 Expected Result 열을 번역"""
    try:
        from openpyxl import load_workbook

        update_status(status="processing", error=None, started_at=datetime.utcnow().isoformat())
        logger.info(f"[{job_id}] Starting translation process for: {input_file.name}")

        wb = load_workbook(input_file)
        ws = wb.active

        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header_row:
            error_msg = "Header row not found"
            logger.error(error_msg)
            update_status(status="error", error=error_msg, completed_at=datetime.utcnow().isoformat())
            return

        steps_col = None
        expected_result_col = None
        for idx, col in enumerate(header_row, 1):
            if col is None:
                continue
            col_lower = str(col).lower().strip()
            if "steps" in col_lower:
                steps_col = idx
                logger.info(f"Found 'Steps' column at index {idx}")
            if "expected" in col_lower and "result" in col_lower:
                expected_result_col = idx
                logger.info(f"Found 'Expected Result' column at index {idx}")

        if not steps_col and not expected_result_col:
            error_msg = "Steps or Expected Result column not found"
            logger.error(error_msg)
            update_status(status="error", error=error_msg, completed_at=datetime.utcnow().isoformat())
            return

        logger.info(
            f"Will translate columns - Steps: {steps_col}, Expected Result: {expected_result_col}"
        )

        def is_translatable(value) -> bool:
            return isinstance(value, str) and value.strip()

        total_cells = 0
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
            if steps_col and steps_col <= len(row) and is_translatable(row[steps_col - 1]):
                total_cells += 1
            if (
                expected_result_col
                and expected_result_col <= len(row)
                and is_translatable(row[expected_result_col - 1])
            ):
                total_cells += 1

        if total_cells == 0:
            error_msg = "No translatable cells found"
            logger.error(error_msg)
            update_status(status="error", error=error_msg, completed_at=datetime.utcnow().isoformat())
            return

        update_status(total=total_cells, current=0, progress=0, estimated_time=0)
        start_time = time.time()

        translated_cells = 0
        for row_idx in range(2, ws.max_row + 1):
            if steps_col:
                cell = ws.cell(row=row_idx, column=steps_col)
                if is_translatable(cell.value):
                    try:
                        translated = translate_with_llm(cell.value, context="Test Steps")
                    except Exception as exc:
                        error_msg = f"{type(exc).__name__}: {exc}"
                        logger.error(f"Translation failed: {error_msg}")
                        set_error_once(error_msg)
                        translated = f"[Translation Error] {cell.value}"
                    cell.value = translated
                    translated_cells += 1

            if expected_result_col:
                cell = ws.cell(row=row_idx, column=expected_result_col)
                if is_translatable(cell.value):
                    try:
                        translated = translate_with_llm(cell.value, context="Expected Result")
                    except Exception as exc:
                        error_msg = f"{type(exc).__name__}: {exc}"
                        logger.error(f"Translation failed: {error_msg}")
                        set_error_once(error_msg)
                        translated = f"[Translation Error] {cell.value}"
                    cell.value = translated
                    translated_cells += 1

            if translated_cells:
                progress = int((translated_cells / total_cells) * 100)
                elapsed = time.time() - start_time
                avg_time_per_cell = elapsed / translated_cells
                remaining_cells = total_cells - translated_cells
                estimated_time = int(avg_time_per_cell * remaining_cells)
                update_status(
                    current=translated_cells,
                    progress=min(progress, 100),
                    estimated_time=max(estimated_time, 0),
                )

        wb.save(output_file)
        wb.close()

        update_status(
            status="completed",
            progress=100,
            estimated_time=0,
            completed_at=datetime.utcnow().isoformat(),
        )
        logger.info(f"[{job_id}] Translation completed: {output_file.name}")

    except Exception as exc:
        import traceback

        error_msg = f"{type(exc).__name__}: {exc}"
        update_status(status="error", error=error_msg, completed_at=datetime.utcnow().isoformat())
        logger.error(f"Error in translation process: {error_msg}")
        logger.debug(traceback.format_exc())
    finally:
        safe_unlink(Path(input_file))

@app.route("/health", methods=["GET"])
def health():
    return jsonify({"status": "ok"})


@app.route("/heartbeat", methods=["POST"])
def heartbeat():
    """브라우저 연결 상태 확인 (주기적 호출)"""
    update_heartbeat()
    return jsonify({"status": "ok", "message": "heartbeat received"})


@app.route("/")
def index():
    return render_template("index.html")


@app.route("/icon.png")
def serve_icon():
    """아이콘 파일 제공"""
    # 우선순위: 1) 루트의 icon.png, 2) templates/icon.png, 3) icon.iconset
    icon_paths = [
        RESOURCE_DIR / "icon.png",
        RESOURCE_DIR / "templates" / "icon.png",
        RESOURCE_DIR / "icon.iconset" / "icon_256x256.png",
        RESOURCE_DIR / "icon.iconset" / "icon_128x128@2x.png",
    ]

    for icon_path in icon_paths:
        if icon_path.exists():
            logger.info(f"Serving icon from: {icon_path}")
            return send_file(str(icon_path), mimetype="image/png")

    logger.error(f"Icon not found in any location. RESOURCE_DIR: {RESOURCE_DIR}")
    return "", 404


@app.route("/upload", methods=["POST"])
def upload_file():
    if is_processing():
        return jsonify({"error": "번역이 진행 중입니다. 완료 후 다시 시도해 주세요."}), 409

    if not get_gemini_model():
        return jsonify({"error": "GEMINI_API_KEY가 설정되지 않았습니다."}), 400

    file = request.files.get("file")
    if not file:
        return jsonify({"error": "파일이 없습니다."}), 400

    if file.filename == "":
        return jsonify({"error": "파일이 선택되지 않았습니다."}), 400

    if not is_allowed_file(file.filename):
        return jsonify({"error": "엑셀 파일(.xlsx)만 업로드 가능합니다."}), 400

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    nonce = uuid4().hex[:8]
    input_filename = secure_filename(f"input_{timestamp}_{nonce}{Path(file.filename).suffix.lower()}")
    output_filename = secure_filename(f"translated_{timestamp}_{nonce}.xlsx")

    input_path = UPLOAD_FOLDER / input_filename
    output_path = OUTPUT_FOLDER / output_filename

    try:
        file.save(str(input_path))
    except Exception as exc:
        logger.error(f"Failed to save upload: {exc}")
        return jsonify({"error": "파일 저장에 실패했습니다."}), 500

    job_id = uuid4().hex
    reset_status(job_id=job_id, output_file=output_filename)
    update_status(status="processing")

    thread = Thread(
        target=process_excel_translation,
        args=(input_path, output_path, job_id),
        daemon=True,
    )
    thread.start()

    return jsonify({"message": "번역이 시작되었습니다.", "filename": output_filename, "job_id": job_id})


@app.route("/status", methods=["GET"])
def get_status():
    return jsonify(get_status_snapshot())


@app.route("/download/<path:filename>", methods=["GET"])
def download_file(filename):
    safe_name = secure_filename(filename)
    if not safe_name or safe_name != filename:
        return jsonify({"error": "잘못된 파일 이름입니다."}), 400

    file_path = OUTPUT_FOLDER / safe_name
    logger.info(f"Download requested: {safe_name}")
    logger.info(f"Looking for file at: {file_path}")

    if file_path.exists():
        logger.info(f"File found, sending: {file_path}")
        return send_from_directory(
            str(OUTPUT_FOLDER), safe_name, as_attachment=True, download_name=safe_name
        )

    logger.error(f"File not found: {file_path}")
    return jsonify({"error": "파일을 찾을 수 없습니다."}), 404


@app.errorhandler(413)
def file_too_large(_error):
    return jsonify({"error": f"파일이 너무 큽니다. (최대 {MAX_UPLOAD_MB}MB)"}), 413


@app.route("/logs", methods=["GET"])
def get_logs():
    """최근 로그를 반환 (웹 UI 콘솔용)"""
    try:
        lines = int(request.args.get('lines', 100))
        lines = min(lines, 1000)  # 최대 1000줄
        
        if not LOG_FILE.exists():
            return jsonify({"logs": [], "message": "로그 파일이 없습니다."})
        
        with open(LOG_FILE, 'r', encoding='utf-8', errors='replace') as f:
            all_lines = f.readlines()
            recent_lines = all_lines[-lines:] if len(all_lines) > lines else all_lines
            
        return jsonify({
            "logs": [line.rstrip() for line in recent_lines],
            "total": len(all_lines),
            "showing": len(recent_lines)
        })
    except Exception as e:
        logger.error(f"Failed to read logs: {e}")
        return jsonify({"error": str(e)}), 500


@app.route("/system-info", methods=["GET"])
def get_system_info():
    """시스템 정보 반환"""
    return jsonify({
        "python_version": sys.version,
        "platform": sys.platform,
        "frozen": getattr(sys, "frozen", False),
        "resource_dir": str(RESOURCE_DIR),
        "data_dir": str(DATA_DIR),
        "log_file": str(LOG_FILE),
        "port": get_server_port()
    })


def get_server_port() -> int:
    env_port = os.environ.get("FLASK_PORT") or os.environ.get("TRANSLATION_SERVER_PORT")
    if env_port:
        try:
            port = int(env_port)
            if 1 <= port <= 65535:
                return port
        except ValueError:
            logger.warning(f"Invalid port in env: {env_port}")
    return DEFAULT_PORT


def is_port_in_use(port: int) -> bool:
    """포트가 사용 중인지 확인"""
    with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
        try:
            s.bind(("127.0.0.1", port))
            return False
        except OSError:
            return True


def kill_process_using_port(port: int) -> bool:
    """특정 포트를 사용 중인 프로세스 종료"""
    try:
        for proc in psutil.process_iter(['pid', 'name', 'connections']):
            try:
                connections = proc.connections()
                for conn in connections:
                    if conn.laddr.port == port and conn.laddr.ip == "127.0.0.1":
                        proc_name = proc.name()
                        proc_pid = proc.pid
                        logger.warning(f"Port {port} is in use by {proc_name} (PID: {proc_pid})")
                        logger.info(f"Terminating process {proc_name} (PID: {proc_pid})...")
                        proc.terminate()
                        proc.wait(timeout=5)
                        logger.info(f"Process terminated successfully")
                        time.sleep(1)  # 포트 해제 대기
                        return True
            except (psutil.NoSuchProcess, psutil.AccessDenied, psutil.TimeoutExpired):
                continue
        return False
    except Exception as e:
        logger.error(f"Failed to kill process on port {port}: {e}")
        return False


def find_available_port(start_port: int, max_attempts: int = 10) -> int:
    """사용 가능한 포트 찾기"""
    for port in range(start_port, start_port + max_attempts):
        if not is_port_in_use(port):
            return port
    return None


def ensure_port_available(port: int, auto_kill: bool = True) -> int:
    """포트가 사용 가능한지 확인하고 필요시 처리"""
    if not is_port_in_use(port):
        logger.info(f"Port {port} is available")
        return port
    
    logger.warning(f"Port {port} is already in use")
    
    if auto_kill:
        # 기존 프로세스 종료 시도
        logger.info("Attempting to terminate the process using the port...")
        if kill_process_using_port(port):
            if not is_port_in_use(port):
                logger.info(f"Port {port} is now available")
                return port
    
    # 다른 포트 찾기
    logger.warning("Looking for an alternative port...")
    alternative_port = find_available_port(port + 1)
    if alternative_port:
        logger.info(f"Using alternative port: {alternative_port}")
        return alternative_port
    
    # 모두 실패
    logger.error(f"Could not find an available port")
    return port  # 일단 시도는 해봄


def open_browser(url: str, delay: float = 1.5):
    """서버 준비 후 브라우저 자동 실행"""
    def _open():
        time.sleep(delay)
        # 서버가 준비될 때까지 최대 10초 대기
        for _ in range(20):
            try:
                urlopen(url, timeout=1)
                break
            except (URLError, Exception):
                time.sleep(0.5)
        
        logger.info(f"Opening browser: {url}")
        try:
            webbrowser.open(url)
        except Exception as e:
            logger.warning(f"Failed to open browser: {e}")
            logger.info(f"Please manually open: {url}")
    
    thread = Thread(target=_open, daemon=True)
    thread.start()


if __name__ == '__main__':
    logger.info("=" * 60)
    logger.info("TestGPT TC Translator Starting")
    logger.info("=" * 60)
    
    # API 키 체크 (실제 초기화는 지연)
    api_key = os.environ.get('GEMINI_API_KEY')
    if api_key:
        logger.info("[OK] GEMINI_API_KEY configured")
    else:
        logger.warning("[Warning] GEMINI_API_KEY not set")
        logger.warning("  Please set API key in the web interface")
    
    # 포트 확인 및 확보
    requested_port = get_server_port()
    available_port = ensure_port_available(requested_port, auto_kill=True)
    
    url = f"http://127.0.0.1:{available_port}"
    
    logger.info(f"Server will run on port {available_port}")
    logger.info(f"Access URL: {url}")
    logger.info("=" * 60)
    
    # Heartbeat 모니터링 시작 (빌드된 exe에서만)
    if getattr(sys, "frozen", False):
        logger.info("Starting heartbeat monitor...")
        heartbeat_thread = Thread(target=check_heartbeat_timeout, daemon=True)
        heartbeat_thread.start()
        
        logger.info("Opening browser automatically...")
        open_browser(url)
    
    # 127.0.0.1로 바인딩하여 localhost 문제 방지
    try:
        app.run(host="127.0.0.1", port=available_port, debug=False, use_reloader=False, threaded=True)
    except OSError as e:
        if "Address already in use" in str(e):
            logger.error(f"Port {available_port} is still in use. Please close other applications and try again.")
            logger.error("Or manually kill the process:")
            logger.error(f"  netstat -ano | findstr :{available_port}")
            logger.error(f"  taskkill /F /PID <PID>")
        else:
            logger.error(f"Failed to start server: {e}")
        sys.exit(1)
