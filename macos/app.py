# -*- coding: utf-8 -*-
import sys
import io
import logging

# Windows 콘솔에서 UTF-8 출력 지원
if sys.platform == 'win32':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

# 로깅 설정
logging.basicConfig(
    level=logging.INFO,  # DEBUG -> INFO로 변경 (로그 줄이기)
    format='[%(levelname)s] %(message)s',
    handlers=[
        logging.StreamHandler(sys.stdout)
    ],
    force=True
)
logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)

from flask import Flask, render_template, request, jsonify, send_file
from flask_cors import CORS
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
import time
from datetime import datetime
import google.generativeai as genai
from threading import Thread
import json

# PyInstaller 빌드 환경을 위한 기본 경로 설정
if getattr(sys, 'frozen', False):
    # PyInstaller로 빌드된 실행 파일인 경우
    BASE_DIR = os.path.dirname(sys.executable)
else:
    # 개발 환경에서 실행되는 경우
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# Flask 앱 생성 (static 폴더를 BASE_DIR로 설정)
app = Flask(__name__, static_folder=BASE_DIR, static_url_path='/static')
CORS(app)

# PyInstaller 빌드 환경을 위한 기본 경로 설정
if getattr(sys, 'frozen', False):
    # PyInstaller로 빌드된 실행 파일인 경우
    # 실행 파일이 있는 디렉토리를 기준으로 설정
    BASE_DIR = os.path.dirname(sys.executable)
else:
    # 개발 환경에서 실행되는 경우
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))

UPLOAD_FOLDER = os.path.join(BASE_DIR, 'uploads')
OUTPUT_FOLDER = os.path.join(BASE_DIR, 'outputs')
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(OUTPUT_FOLDER, exist_ok=True)

logger.info(f"Base Directory: {BASE_DIR}")
logger.info(f"Upload Folder: {UPLOAD_FOLDER}")
logger.info(f"Output Folder: {OUTPUT_FOLDER}")

# 번역 상태를 저장할 전역 변수
translation_status = {
    'progress': 0,
    'total': 0,
    'current': 0,
    'status': 'idle',
    'estimated_time': 0,
    'output_file': None,
    'error': None
}

# Gemini API 설정 (환경변수에서 API 키 가져오기)
gemini_model = None
try:
    api_key = os.environ.get('GEMINI_API_KEY')
    if api_key:
        genai.configure(api_key=api_key)
        gemini_model = genai.GenerativeModel('gemini-2.5-flash')
        logger.info("Gemini API initialized with model: gemini-2.5-flash")
except Exception as e:
    logger.error(f"Failed to initialize Gemini API: {e}")

def translate_with_llm(text, context=""):
    """LLM을 사용하여 테스트 케이스를 전문적으로 번역"""
    if not gemini_model:
        # API 키가 없으면 더미 번역 (테스트용)
        error_msg = "Gemini API model not initialized - check API key"
        logger.error(error_msg)
        return f"[Translation Error] {text}"
    
    if not text or not isinstance(text, str) or not text.strip():
        return text
    
    prompt = f"""You are a senior QA engineer with 30 years of experience in software testing and mobile app testing. 
You are an expert in translating test cases from Korean to English while maintaining technical accuracy and clarity.

Translate the following Korean test case text to English. Keep the translation:
- Professional and technically accurate
- Clear and concise
- Using proper QA/testing terminology
- Maintaining the original meaning and intent
- Preserving line breaks and formatting

{f'Context: {context}' if context else ''}

Korean text to translate:
{text}

Provide ONLY the English translation without any additional explanation or comments."""

    try:
        logger.info(f"Translating text ({len(text)} chars, {len(text.split())} words)")
        response = gemini_model.generate_content(prompt)
        
        if not response or not hasattr(response, 'text'):
            error_msg = "API returned empty or invalid response"
            logger.error(error_msg)
            if translation_status['error'] is None:
                translation_status['error'] = error_msg
            return f"[Translation Error] {text}"
            
        translated = response.text.strip()
        logger.info(f"Translation completed ({len(translated)} chars)")
        return translated
        
    except Exception as e:
        import traceback
        error_msg = f"{type(e).__name__}: {str(e)}"
        logger.error(f"Translation failed - {error_msg}")
        logger.debug(traceback.format_exc())
        
        # 에러를 번역 상태에도 기록
        if translation_status['error'] is None:
            translation_status['error'] = error_msg
        return f"[Translation Error] {text}"

def process_excel_translation(input_file, output_file):
    """엑셀 파일을 읽어 Steps와 Expected Result 열을 번역"""
    global translation_status
    
    try:
        translation_status['status'] = 'processing'
        translation_status['error'] = None
        
        logger.info(f"Starting translation process for: {os.path.basename(input_file)}")
        
        # 엑셀 파일 읽기
        df = pd.read_excel(input_file)
        wb = load_workbook(input_file)
        ws = wb.active
        
        logger.info(f"Loaded Excel file: {len(df)} rows, {len(df.columns)} columns")
        
        # Steps와 Expected Result 열 찾기
        steps_col = None
        expected_result_col = None
        
        for idx, col in enumerate(df.columns, 1):
            col_lower = str(col).lower()
            logger.debug(f"Column {idx}: {col}")
            if 'steps' in col_lower:
                steps_col = idx
                logger.info(f"Found 'Steps' column at index {idx}")
            if 'expected' in col_lower and 'result' in col_lower:
                expected_result_col = idx
                logger.info(f"Found 'Expected Result' column at index {idx}")
        
        if not steps_col and not expected_result_col:
            error_msg = 'Steps or Expected Result column not found'
            logger.error(error_msg)
            translation_status['error'] = error_msg
            translation_status['status'] = 'error'
            return
        
        logger.info(f"Will translate columns - Steps: {steps_col}, Expected Result: {expected_result_col}")
        
        # 번역할 총 셀 수 계산
        rows_to_translate = len(df) - 1  # 헤더 제외
        cols_to_translate = (1 if steps_col else 0) + (1 if expected_result_col else 0)
        total_cells = rows_to_translate * cols_to_translate
        
        translation_status['total'] = total_cells
        translation_status['current'] = 0
        
        logger.info(f"Translation task: {total_cells} cells to translate")
        
        start_time = time.time()
        
        # 각 행 번역
        for row_idx in range(2, len(df) + 2):  # 엑셀은 1-based, 헤더 다음부터
            # Steps 열 번역
            if steps_col:
                cell = ws.cell(row=row_idx, column=steps_col)
                if cell.value and isinstance(cell.value, str) and cell.value.strip():
                    translated = translate_with_llm(cell.value, context="Test Steps")
                    cell.value = translated
                    
                    translation_status['current'] += 1
                    translation_status['progress'] = int((translation_status['current'] / total_cells) * 100)
                    
                    # 예상 소요 시간 계산
                    elapsed = time.time() - start_time
                    if translation_status['current'] > 0:
                        avg_time_per_cell = elapsed / translation_status['current']
                        remaining_cells = total_cells - translation_status['current']
                        translation_status['estimated_time'] = int(avg_time_per_cell * remaining_cells)
            
            # Expected Result 열 번역
            if expected_result_col:
                cell = ws.cell(row=row_idx, column=expected_result_col)
                if cell.value and isinstance(cell.value, str) and cell.value.strip():
                    translated = translate_with_llm(cell.value, context="Expected Result")
                    cell.value = translated
                    
                    translation_status['current'] += 1
                    translation_status['progress'] = int((translation_status['current'] / total_cells) * 100)
                    
                    # 예상 소요 시간 계산
                    elapsed = time.time() - start_time
                    if translation_status['current'] > 0:
                        avg_time_per_cell = elapsed / translation_status['current']
                        remaining_cells = total_cells - translation_status['current']
                        translation_status['estimated_time'] = int(avg_time_per_cell * remaining_cells)
        
        # 번역된 파일 저장
        wb.save(output_file)
        
        translation_status['progress'] = 100
        translation_status['status'] = 'completed'
        translation_status['output_file'] = output_file
        translation_status['estimated_time'] = 0
        
        logger.info(f"Translation completed: {os.path.basename(output_file)}")
        
    except Exception as e:
        error_msg = f"{type(e).__name__}: {str(e)}"
        translation_status['error'] = error_msg
        translation_status['status'] = 'error'
        logger.error(f"Error in translation process: {error_msg}")
        import traceback
        logger.debug(traceback.format_exc())

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/icon.png')
def serve_icon():
    """아이콘 파일 제공"""
    # 우선순위: 1) 루트의 icon.png, 2) icon.iconset의 256x256 아이콘
    icon_paths = [
        os.path.join(BASE_DIR, 'icon.png'),
        os.path.join(BASE_DIR, 'templates', 'icon.png'),
        os.path.join(BASE_DIR, 'icon.iconset', 'icon_256x256.png'),
        os.path.join(BASE_DIR, 'icon.iconset', 'icon_128x128@2x.png')
    ]
    
    for icon_path in icon_paths:
        if os.path.exists(icon_path):
            logger.info(f"Serving icon from: {icon_path}")
            return send_file(icon_path, mimetype='image/png')
    
    logger.error(f"Icon not found in any location. BASE_DIR: {BASE_DIR}")
    return '', 404

@app.route('/upload', methods=['POST'])
def upload_file():
    global translation_status
    
    if 'file' not in request.files:
        return jsonify({'error': '파일이 없습니다.'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': '파일이 선택되지 않았습니다.'}), 400
    
    if not file.filename.endswith(('.xlsx', '.xls')):
        return jsonify({'error': '엑셀 파일만 업로드 가능합니다.'}), 400
    
    # 파일 저장
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    input_filename = f"input_{timestamp}.xlsx"
    output_filename = f"translated_{timestamp}.xlsx"
    
    input_path = os.path.join(UPLOAD_FOLDER, input_filename)
    output_path = os.path.join(OUTPUT_FOLDER, output_filename)
    
    file.save(input_path)
    
    # 번역 상태 초기화
    translation_status = {
        'progress': 0,
        'total': 0,
        'current': 0,
        'status': 'processing',
        'estimated_time': 0,
        'output_file': output_path,
        'error': None
    }
    
    # 백그라운드에서 번역 시작
    thread = Thread(target=process_excel_translation, args=(input_path, output_path))
    thread.start()
    
    return jsonify({
        'message': '번역이 시작되었습니다.',
        'filename': output_filename
    })

@app.route('/status', methods=['GET'])
def get_status():
    return jsonify(translation_status)

@app.route('/download/<filename>', methods=['GET'])
def download_file(filename):
    file_path = os.path.join(OUTPUT_FOLDER, filename)
    logger.info(f"Download requested: {filename}")
    logger.info(f"Looking for file at: {file_path}")
    
    if os.path.exists(file_path):
        logger.info(f"File found, sending: {file_path}")
        return send_file(file_path, as_attachment=True, download_name=filename)
    else:
        logger.error(f"File not found: {file_path}")
        logger.error(f"Output folder contents: {os.listdir(OUTPUT_FOLDER) if os.path.exists(OUTPUT_FOLDER) else 'Folder does not exist'}")
        return jsonify({'error': f'파일을 찾을 수 없습니다: {file_path}'}), 404

if __name__ == '__main__':
    logger.info("=" * 60)
    logger.info("Test Case Translation Tool Starting")
    logger.info("=" * 60)
    if gemini_model:
        logger.info("[OK] Gemini API Connected")
    else:
        logger.warning("[Warning] GEMINI_API_KEY not set")
        logger.warning("  Please register API key in Electron app settings")
    logger.info("Server running on port 5000")
    logger.info("=" * 60)
    
    # Electron 환경에서는 debug=False, reloader=False 사용
    # 127.0.0.1로 바인딩하여 localhost 문제 방지
    app.run(host='127.0.0.1', port=5000, debug=False, use_reloader=False, threaded=True)

